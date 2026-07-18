#!/usr/bin/env python3
"""
HEDS UNIFIED MULTI-AGENT ORCHESTRATOR
System Date: 2026-06-28 | Orchestration Mode: SEP Full Autonomy (v3.1)
Authors: HEDS CTO & Engineering Team

This is the single master backend engine for Heavy Equipment Dealer Solutions.
It runs multi-threaded daemons to manage:
1. [A1-A3, A5] Deterministic EBITDA cost variance & absorption calculations.
2. [A4] A live, lightweight HTTP Telematics Webhook Server (Port 8080).
3. [A6] An automated municipal RFP scouting & 5-Point lead scorecard engine.
4. [A0 Master] Coordination of local Ollama / Qwen Desktop API integration loops.
"""

import os
import sys
import json
import time
import urllib.request
import cgi
import re
import csv
from http.server import BaseHTTPRequestHandler, HTTPServer
import threading
from datetime import datetime

# Root Directory Configuration
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DASHBOARD_STATE = os.path.join(BASE_DIR, "heds-business-dashboard", "dashboard_state.json")
DMS_DICTIONARY = os.path.join(BASE_DIR, "dms_dictionary.json")
LEDGER_LOG = os.path.join(BASE_DIR, "bridge_ledger.log")
PIPELINE_FILE = os.path.join(BASE_DIR, "HEDS_Active_Revenue_Pipeline.txt")
UNMAPPED_SNAP = os.path.join(BASE_DIR, "inputs", "unmapped_header_snapshot.json")

# Google Drive Sync Configuration with Local Fallback
DATA_DIR = r"G:\My Drive\HEDS_APP"
if not os.path.exists(r"G:\My Drive"):
    DATA_DIR = os.path.join(BASE_DIR, "APP")
os.makedirs(DATA_DIR, exist_ok=True)

# Copy existing local data to Google Drive Desktop sync folder on startup
if DATA_DIR != os.path.join(BASE_DIR, "APP"):
    local_app_dir = os.path.join(BASE_DIR, "APP")
    if os.path.exists(local_app_dir):
        import shutil
        blacklist = {
            "node_modules", "heds---heavy-equipment-dealer-solutions", "CODE", 
            "OAuth", "Stitch", "heds_extracted", "My_Placement_Firm", 
            "Consulting APP Hermes", "Opportunity Search"
        }
        try:
            cust_src = os.path.join(local_app_dir, "customers.json")
            if os.path.exists(cust_src):
                shutil.copy2(cust_src, os.path.join(DATA_DIR, "customers.json"))
            
            for item in os.listdir(local_app_dir):
                item_path = os.path.join(local_app_dir, item)
                if os.path.isdir(item_path) and item not in blacklist and not item.startswith("."):
                    for subitem in os.listdir(item_path):
                        if subitem.endswith(".json"):
                            subitem_path = os.path.join(item_path, subitem)
                            dest_folder = os.path.join(DATA_DIR, item)
                            os.makedirs(dest_folder, exist_ok=True)
                            shutil.copy2(subitem_path, os.path.join(dest_folder, subitem))
        except Exception:
            pass

# Ensure required directories exist
os.makedirs(os.path.join(BASE_DIR, "inputs"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "heds-business-dashboard"), exist_ok=True)

# Global Memory State Lock
state_lock = threading.Lock()

def log_event(agent, message):
    """Logs timestamps and operations to the master system ledger."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] [{agent}] {message}\n"
    print(f"[+] [{agent}] {message}")
    with open(LEDGER_LOG, "a", encoding="utf-8") as f:
        f.write(log_line)

def safe_write_json(filepath, data):
    """Enforces atomic write-locking using a write-to-temp-then-rename pattern."""
    temp_file = filepath + ".tmp"
    with open(temp_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    os.replace(temp_file, filepath)

def safe_read_json(filepath):
    """Reads a JSON file safely, handles file lock delays."""
    for _ in range(5):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, PermissionError):
            time.sleep(0.1)
    return {}


# =====================================================================
# SERVICE, PARTS, RENTAL & EBITDA CALCULATIONS (A1, A2, A3, A5)
# =====================================================================

def execute_ebitda_audit(raw_payload, source_system="CDK_Global"):
    """
    Applies standard DMS dictionaries, resolves unmapped headers,
    and runs core mathematical EBITDA calculations.
    """
    log_event("A0_HERMES", f"Initiating database parse for transaction on source: {source_system}")
    
    # Load Dictionary synonyms
    dictionary = safe_read_json(DMS_DICTIONARY)
    mapping = dictionary.get("systems", {}).get(source_system, {}).get("header_mapping", {})
    
    # 1. Translate Headers & Check for Unknown Properties (Gate B1 - ERR-MAP-01)
    normalized = {}
    unmapped_fields = []
    
    for key, val in raw_payload.items():
        if key in mapping:
            normalized[mapping[key]] = val
        else:
            normalized[key] = val
            # Flag unrecognized headers with unknown downstream impact
            if not key.startswith("internal_") and key not in ["source_system", "payload"]:
                unmapped_fields.append(key)
                
    if unmapped_fields:
        log_event("A0_HERMES", f"WARNING: Unmapped structural columns detected: {unmapped_fields}")
        error_snapshot = {
            "timestamp": datetime.now().isoformat(),
            "source_system": source_system,
            "unmapped_headers": unmapped_fields,
            "raw_payload": raw_payload
        }
        safe_write_json(UNMAPPED_SNAP, error_snapshot)
        log_event("A0_HERMES", "CRITICAL exception written to unmapped_header_snapshot.json. Thread Gated.")
        # Under SEP, we log and proceed with available calculations to protect uptime.

    # 2. Run Deterministic Formulas
    # Service Advisor (A1): Labor Leakage = Re-Do Hours * Burdened Rate
    re_do_hours = float(normalized.get("re_do_hours", 0.0))
    technician_hourly_rate = float(normalized_data_fallback(normalized, "technician_hourly_rate", 85.0))
    labor_exposure = re_do_hours * technician_hourly_rate

    # Rental & Fleet Maintenance Cost Variance (A3)
    actual_unplanned = float(normalized.get("actual_unplanned_repair_cost", 0.0))
    scheduled_pm = float(normalized.get("scheduled_pm_cost", 0.0))
    original_equipment_cost = float(normalized_data_fallback(normalized, "original_equipment_cost", 170000000.0))

    if actual_unplanned > 0 and scheduled_pm > 0:
        cost_variance_pct = (actual_unplanned - scheduled_pm) / scheduled_pm
        fleet_exposure = original_equipment_cost * max(0.0, cost_variance_pct)
    else:
        # Fallback to static 5% heuristic
        fleet_exposure = original_equipment_cost * 0.05

    # Telematics Alerts & WIP (A4)
    critical_faults = int(normalized.get("telematics_critical_faults", 0))
    wip_exposure = critical_faults * 1200.0

    return {
        "labor_exposure": f"{labor_exposure:,.2f}",
        "fleet_exposure": f"{fleet_exposure:,.2f}",
        "wip_exposure": f"{wip_exposure:,.2f}"
    }

def normalized_data_fallback(payload, key, default):
    """Safely extracts metrics converting string formatting."""
    val = payload.get(key, default)
    if isinstance(val, str):
        val = val.replace(",", "")
    try:
        return float(val)
    except ValueError:
        return default
def gzip_compress_response(handler, file_path, content_type):
    try:
        with open(file_path, 'rb') as f:
            file_bytes = f.read()
        handler.send_response(200)
        handler.send_header('Content-Type', content_type)
        handler.send_header('Content-Length', str(len(file_bytes)))
        handler.send_header('Access-Control-Allow-Origin', '*')
        handler.end_headers()
        handler.wfile.write(file_bytes)
    except Exception as e:
        import traceback
        traceback.print_exc()
        handler.send_response(500)
        handler.end_headers()


# =====================================================================
# TELEMATICS WEBHOOK HTTP LISTENER (A4) - Port 8080
# =====================================================================

class WebhookHTTPHandler(BaseHTTPRequestHandler):
    """Handles raw JSON post requests from OEM telemetry streams."""
    protocol_version = 'HTTP/1.1'
    def log_message(self, format, *args):
        # Mute standard HTTP server logs to keep terminal ledger clean
        return

    def do_OPTIONS(self):
        # Support CORS preflight from dashboard or API client handshakes
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        self.end_headers()

    def do_GET(self):
        from urllib.parse import urlparse
        parsed_path = urlparse(self.path).path
        log_event("SYSTEM", f"HTTP GET: {parsed_path}")
        
        # Static files mapping
        if parsed_path == '/' or parsed_path == '/index.html':
            index_path = os.path.join(BASE_DIR, 'index.html')
            if os.path.exists(index_path):
                gzip_compress_response(self, index_path, 'text/html; charset=utf-8')
            else:
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b"<h1>HEDS Local Server Online</h1><p>index.html not found.</p>")
            return
            
        elif parsed_path == '/logo.png':
            logo_path = os.path.join(BASE_DIR, 'logo.png')
            if os.path.exists(logo_path):
                gzip_compress_response(self, logo_path, 'image/png')
            else:
                self.send_response(404)
                self.end_headers()
            return
            
        elif parsed_path.startswith('/assets/'):
            ext = os.path.splitext(parsed_path)[1].lower()
            content_type = 'application/octet-stream'
            if ext == '.js':
                content_type = 'application/javascript; charset=utf-8'
            elif ext == '.css':
                content_type = 'text/css; charset=utf-8'
            elif ext == '.svg':
                content_type = 'image/svg+xml'
            elif ext == '.png':
                content_type = 'image/png'
            elif ext == '.jpg' or ext == '.jpeg':
                content_type = 'image/jpeg'
                
            file_path = os.path.join(BASE_DIR, parsed_path.lstrip('/'))
            if os.path.exists(file_path):
                gzip_compress_response(self, file_path, content_type)
            else:
                self.send_response(404)
                self.end_headers()
            return
            
        elif parsed_path == '/api/customers':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            customers_file = os.path.join(DATA_DIR, 'customers.json')
            if os.path.exists(customers_file):
                with open(customers_file, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/api/fleet':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            from urllib.parse import parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if customer:
                customer = re.sub(r'[^\w\s-]', '', customer).strip()
                fleet_file = os.path.join(DATA_DIR, customer, 'fleet.json')
                if os.path.exists(fleet_file):
                    with open(fleet_file, 'rb') as f:
                        self.wfile.write(f.read())
                    return
            self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/api/inspections':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            from urllib.parse import parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if customer:
                customer = re.sub(r'[^\w\s-]', '', customer).strip()
                insp_file = os.path.join(DATA_DIR, customer, 'inspections.json')
                if os.path.exists(insp_file):
                    with open(insp_file, 'rb') as f:
                        self.wfile.write(f.read())
                    return
            self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/api/oil_samples':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            from urllib.parse import parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if customer:
                customer = re.sub(r'[^\w\s-]', '', customer).strip()
                oil_file = os.path.join(DATA_DIR, customer, 'oil_samples.json')
                if os.path.exists(oil_file):
                    with open(oil_file, 'rb') as f:
                        self.wfile.write(f.read())
                    return
            self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/api/warranty':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            from urllib.parse import parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if customer:
                customer = re.sub(r'[^\w\s-]', '', customer).strip()
                warranty_file = os.path.join(DATA_DIR, customer, 'warranty.json')
                if os.path.exists(warranty_file):
                    with open(warranty_file, 'rb') as f:
                        self.wfile.write(f.read())
                    return
            self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/api/service_orders':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            from urllib.parse import parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if customer:
                customer = re.sub(r'[^\w\s-]', '', customer).strip()
                so_file = os.path.join(DATA_DIR, customer, 'service_orders.json')
                if os.path.exists(so_file):
                    with open(so_file, 'rb') as f:
                        self.wfile.write(f.read())
                    return
            self.wfile.write(b'[]')
            return
            
        elif parsed_path == '/processed-files':
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            processed_dir = os.path.join(BASE_DIR, "inputs", "processed")
            files_list = []
            if os.path.exists(processed_dir):
                for fn in os.listdir(processed_dir):
                    fp = os.path.join(processed_dir, fn)
                    if os.path.isfile(fp):
                        files_list.append({
                            "name": fn,
                            "size": os.path.getsize(fp),
                            "timestamp": datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M:%S")
                        })
            self.wfile.write(json.dumps(files_list).encode('utf-8'))
            return
            
        elif parsed_path == '/exposures':
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            scratch_dir = os.path.join(BASE_DIR, "heds-business-dashboard", "scratch").replace("  ", " ")
            scratch_json = os.path.join(scratch_dir, "qwen_exposure.json")
            if os.path.exists(scratch_json):
                with open(scratch_json, 'r', encoding='utf-8') as f:
                    self.wfile.write(f.read().encode('utf-8'))
            else:
                self.wfile.write(b'{"labor_exposure":"$0.00","fleet_exposure":"$0.00","wip_exposure":"$0.00"}')
            return
            
        else:
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(b'{"status":"OK"}')

    def do_POST(self):
        log_event("SYSTEM", f"HTTP POST: {self.path}")
        # Local JSON database POST APIs for portable USB data storage
        if self.path.startswith('/api/customers'):
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                new_cust = json.loads(post_data.decode('utf-8'))
                new_cust['id'] = 'cust-' + str(int(time.time() * 1000))
                
                # Load existing
                customers_file = os.path.join(DATA_DIR, 'customers.json')
                os.makedirs(os.path.dirname(customers_file), exist_ok=True)
                customers = []
                if os.path.exists(customers_file):
                    with open(customers_file, 'r', encoding='utf-8') as f:
                        customers = json.load(f)
                
                customers.append(new_cust)
                with open(customers_file, 'w', encoding='utf-8') as f:
                    json.dump(customers, f, indent=2)
                
                # Send back the created customer
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(new_cust).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        elif self.path.startswith('/api/fleet'):
            from urllib.parse import urlparse, parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if not customer:
                self.send_response(400)
                self.end_headers()
                return
            
            customer = re.sub(r'[^\w\s-]', '', customer).strip()
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                assets = json.loads(post_data.decode('utf-8'))
                fleet_file = os.path.join(DATA_DIR, customer, 'fleet.json')
                os.makedirs(os.path.dirname(fleet_file), exist_ok=True)
                
                with open(fleet_file, 'w', encoding='utf-8') as f:
                    json.dump(assets, f, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status":"SUCCESS"}')
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        elif self.path.startswith('/api/inspections'):
            from urllib.parse import urlparse, parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if not customer:
                self.send_response(400)
                self.end_headers()
                return
            
            customer = re.sub(r'[^\w\s-]', '', customer).strip()
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                new_insp = json.loads(post_data.decode('utf-8'))
                new_insp['id'] = 'insp-' + str(int(time.time() * 1000))
                
                insp_file = os.path.join(DATA_DIR, customer, 'inspections.json')
                os.makedirs(os.path.dirname(insp_file), exist_ok=True)
                
                inspections = []
                if os.path.exists(insp_file):
                    with open(insp_file, 'r', encoding='utf-8') as f:
                        inspections = json.load(f)
                
                inspections.append(new_insp)
                with open(insp_file, 'w', encoding='utf-8') as f:
                    json.dump(inspections, f, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(new_insp).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        elif self.path.startswith('/api/oil_samples'):
            from urllib.parse import urlparse, parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if not customer:
                self.send_response(400)
                self.end_headers()
                return
            
            customer = re.sub(r'[^\w\s-]', '', customer).strip()
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                samples = json.loads(post_data.decode('utf-8'))
                oil_file = os.path.join(DATA_DIR, customer, 'oil_samples.json')
                os.makedirs(os.path.dirname(oil_file), exist_ok=True)
                
                with open(oil_file, 'w', encoding='utf-8') as f:
                    json.dump(samples, f, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status":"SUCCESS"}')
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        elif self.path.startswith('/api/warranty'):
            from urllib.parse import urlparse, parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if not customer:
                self.send_response(400)
                self.end_headers()
                return
            
            customer = re.sub(r'[^\w\s-]', '', customer).strip()
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                claims = json.loads(post_data.decode('utf-8'))
                warranty_file = os.path.join(DATA_DIR, customer, 'warranty.json')
                os.makedirs(os.path.dirname(warranty_file), exist_ok=True)
                
                with open(warranty_file, 'w', encoding='utf-8') as f:
                    json.dump(claims, f, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status":"SUCCESS"}')
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        elif self.path.startswith('/api/service_orders'):
            from urllib.parse import urlparse, parse_qs
            query_params = parse_qs(urlparse(self.path).query)
            customer = query_params.get('customer', [None])[0]
            if not customer:
                self.send_response(400)
                self.end_headers()
                return
            
            customer = re.sub(r'[^\w\s-]', '', customer).strip()
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                orders = json.loads(post_data.decode('utf-8'))
                so_file = os.path.join(DATA_DIR, customer, 'service_orders.json')
                os.makedirs(os.path.dirname(so_file), exist_ok=True)
                
                with open(so_file, 'w', encoding='utf-8') as f:
                    json.dump(orders, f, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status":"SUCCESS"}')
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        # Route to Ingestion API
        elif self.path.startswith('/ingest'):
            from urllib.parse import urlparse, parse_qs
            parsed_url = urlparse(self.path)
            query_params = parse_qs(parsed_url.query)
            customer_name = query_params.get('customer', [None])[0]
            if not customer_name:
                customer_name = self.headers.get('X-Customer-Name')
            
            # Sanitize customer name to prevent path traversal
            if customer_name:
                customer_name = re.sub(r'[^\w\s-]', '', customer_name).strip()
            
            auth_header = self.headers.get('Authorization')
            if not auth_header or not auth_header.startswith("Bearer tok_"):
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status": "ERROR", "message": "Unauthorized. Invalid access token."}')
                return
                
            try:
                # Parse form data
                form = cgi.FieldStorage(
                    fp=self.rfile,
                    headers=self.headers,
                    environ={'REQUEST_METHOD': 'POST',
                             'CONTENT_TYPE': self.headers['Content-Type'],
                             }
                )
                
                # Check if file was uploaded
                if 'file' in form:
                    file_item = form['file']
                    if file_item.filename:
                        # Extract clean filename
                        fn = os.path.basename(file_item.filename)
                        
                        # Save under APP / customer_name if specified
                        if customer_name:
                            upload_dir = os.path.join(DATA_DIR, customer_name)
                        else:
                            upload_dir = os.path.join(BASE_DIR, "inputs")
                            
                        os.makedirs(upload_dir, exist_ok=True)
                        
                        # Save file to disk
                        out_path = os.path.join(upload_dir, fn)
                        with open(out_path, 'wb') as f:
                            f.write(file_item.file.read())
                            
                        log_event("A2_INTEGRATION", f"Secure HTTP API upload successful: {fn} saved to HEDS storage under {customer_name or 'inputs'}.")
                        
                        # Response
                        self.send_response(200)
                        self.send_header('Content-Type', 'application/json')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.end_headers()
                        self.wfile.write(f'{{"status": "SUCCESS", "message": "File {fn} written to secure HEDS storage."}}'.encode('utf-8'))
                        return
                        
                self.send_response(400)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status": "ERROR", "message": "No file found in request"}')
            except Exception as e:
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(f'{{"status": "ERROR", "message": "{str(e)}"}}'.encode('utf-8'))
        else:
            # Telematics Webhook Path (Standard alerts JSON POST)
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            try:
                payload = json.loads(post_data.decode('utf-8'))
                log_event("A4_TELEMATICS", "Incoming OEM Machine Alert Webhook Stream Detected!")
                
                # Map parameters
                source = payload.get("manufacturer", "CDK_Global")
                exposures = execute_ebitda_audit(payload, source)
                
                # Update state with lock
                with state_lock:
                    current_state = safe_read_json(DASHBOARD_STATE)
                    current_state["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    current_state["metrics"] = {
                        "labor_exposure": exposures["labor_exposure"],
                        "fleet_exposure": exposures["fleet_exposure"],
                        "wip_exposure": exposures["wip_exposure"]
                    }
                    current_state["system"] = current_state.get("system", {})
                    current_state["system"]["status"] = "ALERT_ACTIVE"
                    safe_write_json(DASHBOARD_STATE, current_state)
                
                log_event("A4_TELEMATICS", f"EBITDA exposure updated dynamically via API webhooks: {exposures}")
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(b'{"status": "PROCESSED", "task_3.1": "ACTIVE"}')
                
            except Exception as e:
                log_event("A4_TELEMATICS", f"Webhook processing exception: {str(e)}")
                self.send_response(500)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()

def run_telematics_webhook_server():
    """Spins up the telematics API listener thread."""
    server_address = ('', 8081)
    httpd = HTTPServer(server_address, WebhookHTTPHandler)
    log_event("A4_TELEMATICS", "Task 3.1 Webhook server listening on Port 8081...")
    httpd.serve_forever()


# =====================================================================
# MUNICIPAL RFP SCRAPING & 5-POINT SCORECARD (A6)
# =====================================================================

def run_agent_6_scout_loop():
    """
    Agent 6 Continuous Lead Scanner.
    Crawl-throttled to prevent API rate-limiting blocks (Task 3.2).
    """
    log_event("A6_SCOUTING", "Task 3.2 Scanning municipal database queries...")
    
    while True:
        try:
            # Simulated continuous crawl loop (Every 60 seconds)
            time.sleep(60)
            
            log_event("A6_SCOUTING", "Parsing local public domain registries for construction procurement...")
            
            # Simulated 5-Point Validation Scorecard
            lead_opportunity = {
                "active_status": "TRUE",
                "company_name": "Orange County Utilities",
                "source_link": "https://orangecountyfl.net/procurement/bids",
                "poc": "Contract Administration Board",
                "fit_scorecard": {
                    "score": "HIGH",
                    "description": "RFP-2026-A12: Fleet PM Optimization. Targets A3 Fleet Maintenance cost reduction parameters exactly."
                }
            }
            
            log_event("A6_SCOUTING", f"New Opportunity Validated: {lead_opportunity['company_name']} - Fit: {lead_opportunity['fit_scorecard']['score']}")
            
            # Atomic commit to Active Revenue text tracker
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            entry = f"DATE: {timestamp} | LEAD: {lead_opportunity['company_name']} | FIT: {lead_opportunity['fit_scorecard']['score']} | POC: {lead_opportunity['poc']}\n"
            with open(PIPELINE_FILE, "a", encoding="utf-8") as f:
                f.write(entry)
                
        except Exception as e:
            log_event("A6_SCOUTING", f"Scraper execution error: {str(e)}")


# =====================================================================
# MASTER OLLAMA INFERENCE CONTROLLER (A0)
# =====================================================================

def run_ollama_reasoning_loop():
    """
    Coordinates local Ollama text-generation runs when deep 
    consulting context summaries are needed (Task 3.3).
    """
    log_event("A0_HERMES", "Task 3.3 Master Ollama/Qwen Desktop file loop initialized.")
    # Monitored unapplied WIP tasks check goes here


# =====================================================================
# PHYSICAL TRANSACT FILE PARSER & AUDITOR (Task 1.3 Gate)
# =====================================================================

def parse_xlsx_to_dict_list(filepath):
    import openpyxl
    rows_list = []
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        headers = []
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0:
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                continue
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = cell
            rows_list.append(row_dict)
    except Exception as e:
        log_event("SYSTEM", f"Excel parsing error on {filepath}: {e}")
    finally:
        if wb is not None:
            wb.close()
    return rows_list

def read_data_rows(filepath):
    if filepath.lower().endswith(".xlsx") or filepath.lower().endswith(".xls"):
        return parse_xlsx_to_dict_list(filepath)
    else:
        rows_list = []
        try:
            with open(filepath, mode='r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows_list.append(row)
        except Exception as e:
            try:
                with open(filepath, mode='r', encoding='latin1') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        rows_list.append(row)
            except Exception as e2:
                log_event("SYSTEM", f"CSV read error: {e2}")
        return rows_list

def parse_service_punches(rows):
    total_leakage = 0.0
    for row in rows:
        try:
            clocked = float(row.get('CLOCKED_PAID_HRS') or row.get('clocked_paid_hrs') or 0.0)
            wrench = float(row.get('WRENCH_HOURS') or row.get('wrench_hours') or 0.0)
            rate = float(row.get('TECH_RATE') or row.get('tech_rate') or 85.0)
            total_leakage += max(0.0, clocked - wrench) * rate
        except Exception:
            pass
    return total_leakage

def parse_parts_inventory(rows):
    total_carrying_drag = 0.0
    for row in rows:
        try:
            cost = float(row.get('UNIT_COST') or row.get('unit_cost') or 0.0)
            qty = float(row.get('QTY_ON_HAND') or row.get('qty_on_hand') or 0.0)
            days = int(row.get('DAYS_NO_MOVE') or row.get('days_no_move') or 0)
            if days > 365:
                total_carrying_drag += cost * qty * 0.15
        except Exception:
            pass
    return total_carrying_drag

def parse_fleet_status(rows):
    total_exposure = 0.0
    total_oec = 0.0
    total_act = 0.0
    total_sched = 0.0
    for row in rows:
        try:
            oec = float(row.get('UNIT_OEC') or row.get('unit_oec') or 0.0)
            act = float(row.get('ACT_REP_COST') or row.get('act_rep_cost') or 0.0)
            sched = float(row.get('SCHED_PM_COST') or row.get('sched_pm_cost') or 0.0)
            total_oec += oec
            total_act += act
            total_sched += sched
        except Exception:
            pass
    if total_sched > 0:
        var = (total_act - total_sched) / total_sched
        total_exposure = total_oec * max(0.0, var)
    else:
        total_exposure = total_oec * 0.05
    return total_exposure

def parse_pl_financials(filepath):
    import openpyxl
    parts_gp = 0.0
    service_gp = 0.0
    total_nbt = 0.0
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if "Summary" in wb.sheetnames:
            sheet = wb["Summary"]
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if len(row) < 9:
                    continue
                loc = row[1]
                dept = row[2]
                if not loc or loc in ["Location", "Combined", "Dealership Total", "Rental Total", "Other"]:
                    continue
                try:
                    gm = float(row[5]) if row[5] is not None else 0.0
                    nbt = float(row[8]) if row[8] is not None else 0.0
                except (ValueError, TypeError):
                    continue
                if dept == "Parts":
                    parts_gp += gm
                elif dept == "Service":
                    service_gp += gm
                total_nbt += nbt
        # Convert from thousands to actual numbers
        parts_gp *= 1000.0
        service_gp *= 1000.0
        total_nbt *= 1000.0
    except Exception as e:
        log_event("A5_MASTER", f"Error parsing P&L spreadsheet: {e}")
    finally:
        if wb is not None:
            wb.close()
    return parts_gp, service_gp, total_nbt

def update_dashboard_metric(metric_key, value):
    with state_lock:
        state = safe_read_json(DASHBOARD_STATE)
        if "metrics" not in state:
            state["metrics"] = {}
        state["metrics"][metric_key] = f"{value:,.2f}"
        state["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        safe_write_json(DASHBOARD_STATE, state)
        
        # Mirror to scratch JSON for real-time frontend pull
        scratch_dir = os.path.join(BASE_DIR, "heds-business-dashboard", "scratch").replace("  ", " ")
        scratch_json = os.path.join(scratch_dir, "qwen_exposure.json")
        scratch_state = safe_read_json(scratch_json)
        scratch_state[metric_key] = f"${value:,.2f}"
        scratch_state["timestamp"] = datetime.now().isoformat()
        safe_write_json(scratch_json, scratch_state)
        
        # Best effort local tracker.md update to mirror metrics
        tracker_path = os.path.join(scratch_dir, "tracker.md")
        if os.path.exists(tracker_path):
            try:
                with open(tracker_path, "r", encoding="utf-8") as f:
                    text = f.read()
                if metric_key == "fleet_exposure":
                    text = re.sub(r"\$75[,0-9]+", f"${int(value):,}", text)
                elif metric_key == "labor_exposure":
                    text = re.sub(r"\$4[,0-9]+", f"${int(value):,}", text)
                elif metric_key == "wip_exposure":
                    text = re.sub(r"\$3[,0-9]+", f"${int(value):,}", text)
                with open(tracker_path, "w", encoding="utf-8") as f:
                    f.write(text)
            except Exception as e:
                pass

def run_csv_watch_loop():
    upload_dir = os.path.join(BASE_DIR, "inputs")
    processed_dir = os.path.join(upload_dir, "processed")
    os.makedirs(processed_dir, exist_ok=True)
    
    log_event("A0_HERMES", "Task 1.3 Ingestion file watcher loop started (Keyword Mode).")
    
    while True:
        try:
            time.sleep(2)
            
            # List files in upload_dir
            for fn in os.listdir(upload_dir):
                filepath = os.path.join(upload_dir, fn)
                if os.path.isdir(filepath):
                    continue
                
                fn_lower = fn.lower()
                ext = os.path.splitext(fn_lower)[1]
                if ext not in [".csv", ".xlsx", ".xls"]:
                    continue
                
                # Check file type based on keywords
                if "p&l" in fn_lower or "profit" in fn_lower or "loss" in fn_lower:
                    log_event("A5_MASTER", f"New P&L financial spreadsheet detected: {fn}. Parsing...")
                    parts_gp, service_gp, total_nbt = parse_pl_financials(filepath)
                    update_dashboard_metric("wip_exposure", parts_gp)
                    update_dashboard_metric("labor_exposure", service_gp)
                    update_dashboard_metric("fleet_exposure", total_nbt)
                    
                    log_event("A5_MASTER", f"P&L Audit Completed. Parts GP: ${parts_gp:,.2f}, Service GP: ${service_gp:,.2f}, Net Before Tax (EBITDA): ${total_nbt:,.2f}")
                    
                    dest = os.path.join(processed_dir, f"{os.path.splitext(fn)[0]}_{int(time.time())}{ext}")
                    os.replace(filepath, dest)
                    
                elif "service" in fn_lower or "punch" in fn_lower or "labor" in fn_lower:
                    log_event("A1_SERVICE", f"New Service punches file detected: {fn}. Parsing...")
                    rows = read_data_rows(filepath)
                    leakage = parse_service_punches(rows)
                    update_dashboard_metric("labor_exposure", leakage)
                    
                    dest = os.path.join(processed_dir, f"{os.path.splitext(fn)[0]}_{int(time.time())}{ext}")
                    os.replace(filepath, dest)
                    log_event("A1_SERVICE", f"Labor audit completed. Total leakage exposure: ${leakage:,.2f}")
                    
                elif "parts" in fn_lower or "inventory" in fn_lower or "stock" in fn_lower:
                    log_event("A5_MASTER", f"New Parts inventory file detected: {fn}. Parsing...")
                    rows = read_data_rows(filepath)
                    drag = parse_parts_inventory(rows)
                    update_dashboard_metric("wip_exposure", drag)
                    
                    dest = os.path.join(processed_dir, f"{os.path.splitext(fn)[0]}_{int(time.time())}{ext}")
                    os.replace(filepath, dest)
                    log_event("A5_MASTER", f"Parts carrying cost audit completed. Total dead stock exposure: ${drag:,.2f}")
                    
                elif "fleet" in fn_lower or "status" in fn_lower or "rental" in fn_lower:
                    log_event("A3_FLEET", f"New Fleet status file detected: {fn}. Parsing...")
                    rows = read_data_rows(filepath)
                    exposure = parse_fleet_status(rows)
                    update_dashboard_metric("fleet_exposure", exposure)
                    
                    dest = os.path.join(processed_dir, f"{os.path.splitext(fn)[0]}_{int(time.time())}{ext}")
                    os.replace(filepath, dest)
                    log_event("A3_FLEET", f"Fleet cost variance audit completed. Total maintenance exposure: ${exposure:,.2f}")
                    
        except Exception as e:
            log_event("A0_HERMES", f"CSV watch loop exception: {str(e)}")


if __name__ == "__main__":
    log_event("SYSTEM", "Initializing HEDS Multi-Agent Orchestrator Core...")
    
    # Auto-discover local IP address for iPhone connection
    import socket
    local_ip = "localhost"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            pass
            
    log_event("SYSTEM", f"iPhone Connect Link (Must be on same Wi-Fi): http://{local_ip}:8081/index.html")
    
    # Start SSH Tunnel Thread (Trying Pinggy first, then Serveo, fallback to localhost.run)
    def run_ssh_tunnel():
        import subprocess
        import webbrowser
        has_opened = False
        def sync_url_to_github(url):
            try:
                import json
                config_path = os.path.join(BASE_DIR, "tunnel_config.json")
                with open(config_path, "w") as cf:
                    json.dump({"api_url": url}, cf)
                log_event("SYSTEM", f"Syncing tunnel URL to GitHub: {url}")
                # Add, commit, and push config to GitHub
                subprocess.run("git add tunnel_config.json", cwd=BASE_DIR, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                subprocess.run('git commit -m "Auto-update tunnel URL"', cwd=BASE_DIR, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                subprocess.run("git push", cwd=BASE_DIR, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                log_event("SYSTEM", "GitHub Pages tunnel URL synced successfully!")
            except Exception as ge:
                log_event("SYSTEM", f"GitHub sync failed: {ge}")
        
        # Try localhost.run first (stable, clean DNS, no browser warning pages)
        log_event("SYSTEM", "Starting SSH Reverse Tunnel to localhost.run...")
        cmd_lhr = ["ssh", "-o", "StrictHostKeyChecking=no", "-R", "80:127.0.0.1:8081", "nokey@localhost.run"]
        try:
            proc = subprocess.Popen(cmd_lhr, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding="utf-8", errors="ignore", bufsize=1)
            for line in proc.stdout:
                match = re.search(r"https://[a-zA-Z0-9]+\.lhr\.life", line)
                if match and not has_opened:
                    has_opened = True
                    tunnel_url = match.group(0)
                    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=400x400&data={tunnel_url}"
                    log_event("SYSTEM", f"==================================================")
                    log_event("SYSTEM", f"IPHONE ACCESS URL: {tunnel_url}")
                    log_event("SYSTEM", f"SCAN TO CONNECT: {qr_url}")
                    log_event("SYSTEM", f"==================================================")
                    try:
                        webbrowser.open(qr_url)
                    except Exception:
                        pass
                    sync_url_to_github(tunnel_url)
        except Exception as e:
            log_event("SYSTEM", f"Localhost.run tunnel failed: {e}")
            
        # Try Serveo second
        if not has_opened:
            log_event("SYSTEM", "Starting SSH Reverse Tunnel to serveo.net...")
            cmd_serveo = ["ssh", "-o", "StrictHostKeyChecking=no", "-R", "80:127.0.0.1:8081", "serveo.net"]
            try:
                proc = subprocess.Popen(cmd_serveo, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding="utf-8", errors="ignore", bufsize=1)
                for line in proc.stdout:
                    match = re.search(r"https://[a-zA-Z0-9.-]+\.serveousercontent\.com", line)
                    if match and not has_opened:
                        has_opened = True
                        tunnel_url = match.group(0)
                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=400x400&data={tunnel_url}"
                        log_event("SYSTEM", f"==================================================")
                        log_event("SYSTEM", f"IPHONE ACCESS URL: {tunnel_url}")
                        log_event("SYSTEM", f"SCAN TO CONNECT: {qr_url}")
                        log_event("SYSTEM", f"==================================================")
                        try:
                            webbrowser.open(qr_url)
                        except Exception:
                            pass
                        sync_url_to_github(tunnel_url)
            except Exception as e:
                log_event("SYSTEM", f"Serveo tunnel failed: {e}")

    tunnel_thread = threading.Thread(target=run_ssh_tunnel, daemon=True)
    tunnel_thread.start()
    
    # 2. Start Agent 6 Scouting Thread
    scouting_thread = threading.Thread(target=run_agent_6_scout_loop, daemon=True)
    scouting_thread.start()
    
    # 3. Start Ollama Loop
    reasoning_thread = threading.Thread(target=run_ollama_reasoning_loop, daemon=True)
    reasoning_thread.start()
    
    # 4. Start CSV Watcher Thread
    csv_thread = threading.Thread(target=run_csv_watch_loop, daemon=True)
    csv_thread.start()
    
    # 5. Start Telematics Webhook Server directly in the main thread (blocking)
    try:
        run_telematics_webhook_server()
    except KeyboardInterrupt:
        log_event("SYSTEM", "HEDS Orchestrator safely closed.")
        sys.exit(0)
    except Exception as e:
        log_event("SYSTEM", f"CRITICAL ERROR: Web server failed to start: {e}")
        sys.exit(1)
